import math
import pandas as pd
import numpy as np
import argparse
import openpyxl
from openpyxl.styles import PatternFill


class Params:
    def __init__(self, pos=0, speed=0, reaction_time=0, accel=0, length=0):
        self.pos = pos
        self.speed = speed
        self.reaction_time = reaction_time
        self.accel = accel
        self.length = length


def calculate_collision_with_rss(ego_params: Params, front_vehicle_params: Params, rear_vehicle_params: Params):
    ego_pos, ego_speed, rt, ego_acc, ego_length = ego_params.pos, ego_params.speed, ego_params.reaction_time, ego_params.accel, ego_params.length

    front_pos, front_speed, _, front_acc, front_length = front_vehicle_params.pos, front_vehicle_params.speed, front_vehicle_params.reaction_time, front_vehicle_params.accel, front_vehicle_params.length
    is_front_collision = front_pos - ego_pos - front_length / 2 < ego_speed * rt + \
        (ego_speed ** 2 / (-2.0 * ego_acc)) - (front_speed ** 2) / (-2 * front_acc) + ego_length / 2

    rear_pos, rear_speed, rt, rear_acc, rear_length = rear_vehicle_params.pos, rear_vehicle_params.speed, rear_vehicle_params.reaction_time, rear_vehicle_params.accel, rear_vehicle_params.length
    is_rear_collision = ego_pos - rear_pos - rear_length / 2 < rear_speed * rt + \
        (rear_speed ** 2 / (-2.0 * rear_acc)) - (ego_speed ** 2) / (-2 * ego_acc) + ego_length / 2

    return is_front_collision, is_rear_collision


def save_to_excel_with_colors_and_params(collision_table, file_name, ego_params, front_vehicle_params, rear_vehicle_params, use_same_speed, use_same_accel):
    writer = pd.ExcelWriter(file_name, engine='openpyxl')
    collision_table.to_excel(writer, sheet_name='Collision Table')

    workbook = writer.book
    worksheet = writer.sheets['Collision Table']

    # Add header for units
    worksheet['A1'] = 'Acceleration (m/s²) / Speed (m/s)'

    # Coloring the cells based on collision type
    color_mapping = {
        'None': '00FF00',  # Green
        'Rear': '0000FF',  # Blue
        'Front': 'FF0000',  # Red
        'Both': '800080'   # Purple
    }

    for row in worksheet.iter_rows(min_row=0, min_col=0, max_row=worksheet.max_row, max_col=worksheet.max_column):
        for cell in row:
            collision_type = cell.value
            if collision_type in color_mapping:
                cell.fill = PatternFill(
                    start_color=color_mapping[collision_type], end_color=color_mapping[collision_type], fill_type='solid')

    # Add vehicle parameters below the table
    params = [
        ('Ego Vehicle', ego_params),
        ('Front Vehicle', front_vehicle_params),
        ('Rear Vehicle', rear_vehicle_params)
    ]

    start_row = worksheet.max_row + 2
    for label, param in params:
        worksheet.cell(row=start_row, column=1, value=label)
        worksheet.cell(row=start_row + 1, column=1, value='Position (m)')
        worksheet.cell(row=start_row + 1, column=2, value=param.pos)
        worksheet.cell(row=start_row + 2, column=1, value='Speed (m/s)')
        worksheet.cell(row=start_row + 2, column=2,
                       value='same as ego vehicle' if use_same_speed else param.speed)
        worksheet.cell(row=start_row + 3, column=1, value='Reaction Time (s)')
        worksheet.cell(row=start_row + 3, column=2, value=param.reaction_time)
        worksheet.cell(row=start_row + 4, column=1, value='Acceleration (m/s²)')
        worksheet.cell(row=start_row + 4, column=2,
                       value='same as ego vehicle' if use_same_accel else param.accel)
        worksheet.cell(row=start_row + 5, column=1, value='Length (m)')
        worksheet.cell(row=start_row + 5, column=2, value=param.length)
        start_row += 7

    writer.save()


def main(args):
    front_vehicle_params = Params(pos=30, speed=10,
                                  reaction_time=1.0, accel=-3.0, length=5)
    rear_vehicle_params = Params(pos=-30, speed=10,
                                 reaction_time=1.0, accel=-3.0, length=5)

    speeds = np.arange(1, 31)  # 1 m/s to 30 m/s
    accelerations = np.linspace(-1.0, -10, 19)  # -1.0 m/s² to -10 m/s²

    collision_table = pd.DataFrame(
        index=[f"{a:.1f} m/s²" for a in accelerations], columns=[f"{s} m/s" for s in speeds])

    for speed in speeds:
        for accel in accelerations:
            if args.use_same_speed:
                front_vehicle_params.speed = speed
                rear_vehicle_params.speed = speed

            if args.use_same_acceleration:
                front_vehicle_params.accel = accel
                rear_vehicle_params.accel = accel

            ego_params = Params(pos=0, speed=speed, reaction_time=1.5, accel=accel, length=5)
            is_front_collision, is_rear_collision = calculate_collision_with_rss(
                ego_params, front_vehicle_params, rear_vehicle_params)
            if is_front_collision and is_rear_collision:
                collision_table.at[f"{accel:.1f} m/s²", f"{speed} m/s"] = "Both"
            elif is_front_collision:
                collision_table.at[f"{accel:.1f} m/s²", f"{speed} m/s"] = "Front"
            elif is_rear_collision:
                collision_table.at[f"{accel:.1f} m/s²", f"{speed} m/s"] = "Rear"
            else:
                collision_table.at[f"{accel:.1f} m/s²", f"{speed} m/s"] = "None"

    if args.output_file:
        save_to_excel_with_colors_and_params(collision_table, args.output_file, ego_params,
                                             front_vehicle_params, rear_vehicle_params, args.use_same_speed, args.use_same_acceleration)
        print(f"Collision table saved to {args.output_file}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Calculate collision table using RSS model.")
    parser.add_argument('--use-same-speed', action='store_true',
                        help='Use the same speed for front and rear vehicles as the ego vehicle.')
    parser.add_argument('--use-same-acceleration', action='store_true',
                        help='Use the same acceleration for front and rear vehicles as the ego vehicle.')
    parser.add_argument('--output-file', type=str,
                        help='Output file path for the collision table (Excel format).')
    args = parser.parse_args()
    main(args)
